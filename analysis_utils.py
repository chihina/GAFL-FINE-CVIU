import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import numpy as np
import os
import sys
import json
import shutil
import glob

def get_mode(analyze_name, model_name):
    if 'prev' in analyze_name:
        mode = 'people'
    else:
        mode = 'scene'
    
    return mode

def get_seed_list(model_exp_name, saved_result_dir):
    model_exp_name_search_list = get_seed_exp(model_exp_name, saved_result_dir)
    # for i in sorted(model_exp_name_search_list):
        # print(i)
    # sys.stdout.flush()
    seed_list = [os.path.basename(model_exp_name).split('sd')[1].split('_')[0] for model_exp_name in model_exp_name_search_list]
    return seed_list

def get_seed_exp(model_exp_name, saved_result_dir):
    if 'sd' in model_exp_name:
        model_exp_name_pre, model_exp_name_post = model_exp_name.split('sd')[0], model_exp_name.split('sd')[1][2:].split('<')[0]
        model_exp_name_pre = glob.escape(model_exp_name_pre)
        model_exp_name_post = glob.escape(model_exp_name_post)
        model_exp_name_search = f'{model_exp_name_pre}sd*{model_exp_name_post}*'
        model_exp_name_search_list_child = glob.glob(os.path.join(saved_result_dir, model_exp_name_search, 'eval_gar_metrics_*'))
        model_exp_name_search_list = [os.path.dirname(model_exp_name_search_file) for model_exp_name_search_file in model_exp_name_search_list_child]

        # use_seed_list = [50]
        # use_seed_list = list(range(50, 56))
        use_seed_list = list(range(50, 60))
        model_exp_name_search_list_update = []
        found_seed_list = []
        for model_exp_name in model_exp_name_search_list:
            for seed_num in use_seed_list:
                if f'sd{seed_num}' in model_exp_name and f'sd{seed_num}' not in found_seed_list:
                    model_exp_name_search_list_update.append(model_exp_name)
                    found_seed_list.append(f'sd{seed_num}')
                    break
        model_exp_name_search_list = model_exp_name_search_list_update
    else:
        model_exp_name_search_list = [os.path.join(saved_result_dir, model_exp_name)]

    model_exp_name_search_list = list(set(model_exp_name_search_list))
    return model_exp_name_search_list

def transform_col_names(df):
    ret_num_set = set([int(col.split('@')[1].split(' ')[0]) for col in df.columns])
    ret_num_max = max(ret_num_set)
    col_names = df.columns
    col_names_transformed = [i.replace(f'@{ret_num_max}', '@ALL') for i in col_names]
    df.columns = col_names_transformed
    return df

def modify_multi_df(df_eval_results, model_exp_name_list, model_name_list_multi, saved_result_dir, use_query_type, is_paper_flag, use_mode):
    for model_exp_name_idx, model_exp_name in enumerate(model_exp_name_list):
        analyze_name_list = []
        model_name_list = []
        model_exp_name_search_list = get_seed_exp(model_exp_name, saved_result_dir)
        for model_exp_name_search in model_exp_name_search_list:
            analyze_name_list.append(os.path.basename(model_exp_name_search))
            model_name_list.append(model_name_list_multi[model_exp_name_idx])
        df_eval_results_ours = write_excel(analyze_name_list, model_name_list, saved_result_dir, use_query_type, use_mode, is_paper=is_paper_flag)

        # save results in each seed for detailed analysis
        model_type = 'wokp' if 'w/o L-kp' in model_name_list_multi[0] else 'wkp'
        seed_list = get_seed_list(model_exp_name, saved_result_dir)
        df_eval_results_ours.index = seed_list
        df_eval_results_ours = df_eval_results_ours.sort_index()
        save_excel_dir = os.path.join('analysis', 'excel_details')
        if not os.path.exists(save_excel_dir):
            os.makedirs(save_excel_dir)
        save_excel_file_path = os.path.join(save_excel_dir, f'{model_type}_{use_query_type}.xlsx')
        df_eval_results_ours.to_excel(save_excel_file_path, sheet_name='all')
        
        df_eval_results_ours_mean = pd.DataFrame(df_eval_results_ours.mean(axis=0)).T
        trial_num = len(analyze_name_list)
        if trial_num > 1:
            trial_print = f' ({trial_num} trials)'
        else:
            trial_print = f' ({trial_num} trial)'
        df_eval_results_ours_mean.index = [f'{model_name_list_multi[model_exp_name_idx]}{trial_print}']
        df_eval_results_ours_mean = transform_col_names(df_eval_results_ours_mean)
        df_eval_results = pd.concat([df_eval_results, df_eval_results_ours_mean], axis=0)
    
    return df_eval_results

def intergrate_gafl(df_eval_results, analyze_name_list, model_name_list, saved_result_dir, use_query_type, is_paper_flag, use_mode, seed_list=None):
    df_seed_all = pd.DataFrame()
    if seed_list is None:
        df_seed_mean = write_excel(analyze_name_list, model_name_list, saved_result_dir, use_query_type, use_mode, is_paper=is_paper_flag, is_seed=False)
        df_seed_mean = transform_col_names(df_seed_mean)
    else:
        for seed_num in seed_list:
            seed_id = f'_sd{seed_num}'
            df_seed = write_excel(analyze_name_list, model_name_list, saved_result_dir, use_query_type, use_mode, is_paper=is_paper_flag, seed_id=seed_id)
            df_seed = transform_col_names(df_seed)
            df_seed_all = pd.concat([df_seed_all, df_seed], axis=0)
        
        # sort df_eval_results_ours based on the index
        df_seed_all.index = seed_list
        df_seed_all = df_seed_all.sort_index()
        save_excel_dir = os.path.join('analysis', 'excel_details')
        if not os.path.exists(save_excel_dir):
            os.makedirs(save_excel_dir)
        save_excel_file_path = os.path.join(save_excel_dir, f'gafl_{use_query_type}.xlsx')
        df_seed_all.to_excel(save_excel_file_path, sheet_name='all')
        df_seed_mean = pd.DataFrame(df_seed_all.mean(axis=0)).T
        df_seed_mean.index = [f'{model_name_list[0]}']

    return df_seed_mean

def save_avg_results(result_type, df_eval_results_all_grp, use_mode, use_dataset_name, saved_result_analysis_dir, is_paper_flag, exe_sign):
    save_excel_file_name = f'{exe_sign}_{use_dataset_name}_{use_mode}_{result_type}.xlsx'
    save_excel_file_path = os.path.join(saved_result_analysis_dir, save_excel_file_name)
    
    precision_1_values = df_eval_results_all_grp['Precision@10 local']
    # precision_1_values = df_eval_results_all_grp['Precision@10 local middle']
    df_eval_results_all_grp = df_eval_results_all_grp.loc[precision_1_values.sort_values(ascending=True).index]
    df_eval_results_all_grp.to_excel(save_excel_file_path, sheet_name='all')
    refine_excel(save_excel_file_path, is_paper=is_paper_flag)

def save_detail_results(df_eval_results_all, use_mode, use_dataset_name, saved_result_analysis_dir, use_query_type_list, is_paper_flag, exe_sign, use_key='Precision@1'):
    # write excel for each query type
    df_all = pd.concat(df_eval_results_all, axis=0)

    # calculate MPCA for each query type
    df_eval_results_all_grp_mpca = df_all.groupby(df_all.index).mean()

    exp_type_num = df_eval_results_all_grp_mpca.shape[0]
    df_query_detail = pd.DataFrame()
    for use_query_type_idx, use_query_type in enumerate(use_query_type_list):
        print(f'Query type: {use_query_type}')
        df_eval_results_query = df_all.iloc[use_query_type_idx*exp_type_num:(use_query_type_idx+1)*exp_type_num]

        save_excel_file_name = f'{exe_sign}_{use_dataset_name}_{use_mode}_{use_query_type}.xlsx'
        save_excel_file_path = os.path.join(saved_result_analysis_dir, save_excel_file_name)
        df_eval_results_query.to_excel(save_excel_file_path, sheet_name='all')
        refine_excel(save_excel_file_path, is_paper=is_paper_flag)
        
        if not use_key in df_eval_results_query.columns:
            print(f'No {use_key} in the columns')
            return
        
        df_eval_results_query_key = df_eval_results_query[use_key]
        df_query_detail = pd.concat([df_query_detail, df_eval_results_query_key], axis=1)
    df_query_detail.columns = use_query_type_list

    # save df_query_detail
    save_excel_file_name_query = f'{exe_sign}_{use_dataset_name}_{use_mode}_details_{use_key}.xlsx'
    save_excel_file_path_query = os.path.join(saved_result_analysis_dir, save_excel_file_name_query)
    df_query_detail.to_excel(save_excel_file_path_query, sheet_name='all')
    
    # refine excel
    refine_excel(save_excel_file_path_query, is_paper=is_paper_flag)

def analyze_results(df_eval_results_all, use_query_type_list, use_mode, use_dataset_name, saved_result_analysis_dir, is_paper_flag, exe_sign):
    df_all = pd.concat(df_eval_results_all, axis=0)

    # calculate MPCA for each query type
    df_eval_results_all_grp_mpca = df_all.groupby(df_all.index).mean()

    # calculate MCA for each query type
    query_stats_dir = os.path.join('analysis', 'query_stats')
    activity_weights = get_activity_weight(query_stats_dir, use_query_type_list, use_dataset_name)
    df_all = pd.concat(df_eval_results_all, axis=0)
    model_num = int(df_all.shape[0] / len(use_query_type_list))
    for activity_idx, activity in enumerate(use_query_type_list):
        action_idx = use_query_type_list.index(activity)
        df_all.iloc[action_idx*model_num:(action_idx+1)*model_num] *= activity_weights[activity]
    df_eval_results_all_grp_mca = df_all.groupby(df_all.index).sum()

    # save mca and mpca results
    result_types = ['mca', 'mpca']
    result_dic = {'mca': df_eval_results_all_grp_mca, 'mpca': df_eval_results_all_grp_mpca}
    for result_type, df_eval_results_all_grp in result_dic.items():
        save_avg_results(result_type, df_eval_results_all_grp, use_mode, use_dataset_name, saved_result_analysis_dir, is_paper_flag, exe_sign)

    # save detail results
    use_key_list = [
                    # 'Precision@1', 
                    # 'Precision@1 local',
                    # 'Precision@1 middle', 
                    # 'Precision@1 local middle', 
                    # 'Precision@5', 
                    'Precision@10 local',
                    # 'Precision@5 middle', 
                    # 'Precision@5 local middle', 
                    # 'Precision@ALL local',
                    # 'Precision@ALL',
    ]
    for use_key in use_key_list:
        save_detail_results(df_eval_results_all, use_mode, use_dataset_name, saved_result_analysis_dir, use_query_type_list, is_paper_flag, exe_sign, use_key=use_key)

def write_excel(analyze_name_list, model_name_list, saved_result_dir, query_type, use_mode, is_paper=False, is_seed=True, seed_id=''):
    eval_results_list = []

    for analyze_name, model_name in zip(analyze_name_list, model_name_list):
        if use_mode == 'people_scene':
            use_mode_mod = get_mode(analyze_name, model_name)
        else:
            use_mode_mod = use_mode
        json_file_path_gar = os.path.join(saved_result_dir, analyze_name, f'eval_gar_metrics_{query_type}{seed_id}.json')
        with open(json_file_path_gar, 'r') as f:
            eval_results_dic_gar = json.load(f)
        eval_results_dic = {**eval_results_dic_gar}

        # json_file_path_sampling = os.path.join(saved_result_dir, analyze_name, f'sampling.json')
        # with open(json_file_path_sampling, 'r') as f:
        #     eval_results_dic_sampling = json.load(f)
        # selected_target_ratio = eval_results_dic_sampling['non_query_indices_mean']
        # non_query_indices = eval_results_dic_sampling['non_query_indices']
        # selected_sample_num = len(non_query_indices)
        # eval_results_dic['selected_target_ratio'] = selected_target_ratio
        # eval_results_dic['selected_sample_num'] = selected_sample_num

        eval_results_dic_update = update_eval_dic(eval_results_dic, use_mode_mod, is_seed=is_seed)
        eval_results_list.append(list(eval_results_dic_update.values()))
        eval_metrics_list = list(eval_results_dic_update.keys())
    eval_results_array = np.array(eval_results_list)
    df_eval_results = pd.DataFrame(eval_results_array, model_name_list, eval_metrics_list)
    
    return df_eval_results

def refine_excel(save_excel_file_path, is_paper=False):
    wb = openpyxl.load_workbook(save_excel_file_path)
    ws = wb['all']

    col_names = [cell.value for cell in ws[1]]
    for col_idx, col in enumerate(ws.iter_cols()):
        if col_idx != 0:
            for row_idx, cell in enumerate(col):
                if row_idx != 0:
                    col_name = col_names[col_idx]
                    cell.value = round(float(cell.value), 3)

                    if 'gafl' in save_excel_file_path:
                        cell.number_format = '0.0'
                    else:
                        cell.number_format = '0.000'

    for col_idx, col in enumerate(ws.iter_cols()):
        col_name = col_names[col_idx]
        if col_idx != 0:
            row_val_list = [cell.value for cell in col if type(cell.value) is not str]
            row_val_max = sorted(row_val_list, reverse=True)[0]
            row_val_second_max = sorted(row_val_list, reverse=True)[1] if len(row_val_list) > 1 else row_val_max

            for row_idx, cell in enumerate(col):
                if cell.value == row_val_max:
                    cell.fill = PatternFill(fgColor='FFFF00', bgColor="FFFF00", fill_type = "solid")
                    if is_paper:
                        if 'gafl' in save_excel_file_path:
                            cell.value = r'\red' + '{' + f'{cell.value:.1f}' + '}'
                        else:
                            cell.value = r'\red' + '{' + f'{cell.value:.3f}' + '}'

                # visualize the second-best values
                # if cell.value == row_val_second_max:
                #     cell.fill = PatternFill(fgColor='00FFFF', bgColor="00FFFF", fill_type = "solid")
                #     if is_paper:
                #         if 'gafl' in save_excel_file_path:
                #             cell.value = r'\blue' + '{' + f'{cell.value:.1f}' + '}'
                #         else:
                #             cell.value = r'\blue' + '{' + f'{cell.value:.3f}' + '}'

    wb.save(save_excel_file_path)

def update_eval_dic(eval_results_dic, mode, is_seed=True):
    eval_results_dic_update = {}
    # selected_target_ratio = eval_results_dic['selected_target_ratio']
    # selected_sample_num = eval_results_dic['selected_sample_num']
    eval_results_dic_key = list(eval_results_dic.keys())
    eval_results_dic_key_hit = [key.split('hit@')[1].split()[0] for key in eval_results_dic_key if 'hit' in key]

    hit_idx_start = 1
    hit_idx_end = 10
    search_hit_indices = [i for i in range(hit_idx_start, hit_idx_end+1)]
    for i in range(10, 51, 10):
    # for i in range(10, 51, 20):
        search_hit_indices.append(i)
    hit_idx_max_class = max([int(hit_idx) for hit_idx in eval_results_dic_key_hit])
    search_hit_indices.append(hit_idx_max_class)

    base_key = 'Query retrieval accuracy hit@'

    for query_type in [' query']:
        for agg_type in [' sum', '']:
            agg_type_save = '(sum)' if agg_type == ' sum' else '(normal)'
            target_data_list = [' user query', ''] if is_seed else ['']
            # target_data_list = [' user query fused (max)', ' user query', ''] if is_seed else ['']
            # target_data_list = [' user query fused (max)', ' user query fused (mean)', ' user query', ''] if is_seed else ['']
            for target_data in target_data_list:
                if target_data == ' user query':
                    target_data_save = ' local'
                elif target_data == ' user query fused (max)':
                    target_data_save = ' total-mx'
                elif target_data == ' user query fused (mean)':
                    target_data_save = ' total-mn'
                elif target_data == ' user query fused':
                    target_data_save = ' total'
                else:
                    target_data_save = ''

                # db_type_list = [' wofns', ''] if is_seed else ['']
                # db_type_list = [' wofns'] if is_seed else ['']
                db_type_list = ['']
                for db_type in db_type_list:
                    db_data_save = ' middle' if db_type == ' wofns' else ''
                    for hit_idx in search_hit_indices:
                        search_key = f'{base_key}{hit_idx}{agg_type}{db_type}{target_data}{query_type} ({mode})'
                        search_value = eval_results_dic[search_key]
                        # selected_target_num = int(selected_target_ratio*selected_sample_num)

                        if agg_type_save == '(sum)':
                            # eval_results_dic_update[f'Precision@{save_key}'] = search_value / hit_idx
                            # search_value_update = search_value + selected_target_num
                            # hit_idx_update = hit_idx + selected_sample_num
                            # save_key = f'{hit_idx_update}{target_data_save}{db_data_save}'
                            # eval_results_dic_update[f'Precision@{save_key}'] = search_value_update / hit_idx_update
                            save_key = f'{hit_idx}{target_data_save}{db_data_save}'
                            eval_results_dic_update[f'Precision@{save_key}'] = search_value / hit_idx
                        elif agg_type_save == '(normal)':
                            save_key = f'{hit_idx}{target_data_save}{db_data_save}'
                            # search_value_update = 1 if selected_target_num > 0 else search_value
                            eval_results_dic_update[f'Hit@{save_key}'] = search_value
                            # eval_results_dic_update[f'Hit@{save_key}'] = search_value_update

    # eval_results_dic_update = {}
    # for hit_idx in search_hit_indices:
    #     search_key = f'{base_key}{hit_idx} sum query wofns ({mode})'
    #     search_value = eval_results_dic[search_key]
    #     save_key = f'{hit_idx}'
    #     eval_results_dic_update[f'Precision@{save_key}'] = search_value / hit_idx

    return eval_results_dic_update

def collect_visualization_results(save_vis_dir, analyze_name_list, model_name_list, saved_result_dir):
    """_summary_
    Args:
        save_vis_dir (str): Directory for saving visualization results
        analyze_name_list (list): List of analyze names
        model_name_list (list): List of model names for our understanding
        saved_result_dir (str): Direcoty in which results are saved
    """

    if not os.path.exists(save_vis_dir):
        os.makedirs(save_vis_dir)
    for analyze_name, model_name in zip(analyze_name_list, model_name_list):
        saved_result_analysis_dir = os.path.join(saved_result_dir, analyze_name)
        for mode in ['people', 'scene']:
            if 'CAD' in analyze_name:
                mode = 'people'
            else:
                if not 'ours' in analyze_name:
                    mode = 'people'
                else:
                    mode = 'scene'

            cm_file_path = os.path.join(saved_result_analysis_dir, f'confusion_matrix_gar_{mode}.png')
            cm_save_path = os.path.join(save_vis_dir, f'cm_{model_name}_{mode}.png')
            shutil.copy(cm_file_path, cm_save_path)
            tsne_file_wo_legend_path = os.path.join(saved_result_analysis_dir, f'tsne_{mode}_wo_legend.png')
            tsne_save_wo_legend_path = os.path.join(save_vis_dir, f'tsne_{model_name}_{mode}_wo_legend.png')
            shutil.copy(tsne_file_wo_legend_path, tsne_save_wo_legend_path)
            tsne_file_path = os.path.join(saved_result_analysis_dir, f'tsne_{mode}.png')
            tsne_save_path = os.path.join(save_vis_dir, f'tsne_legend.png')
            shutil.copy(tsne_file_path, tsne_save_path)

def collect_nn_results(save_nn_dir, analyze_name_list, model_name_list, saved_result_dir):
    """_summary_
    Args:
        save_nn_dir (str): Directory for saving visualization results
        analyze_name_list (list): List of analyze names
        model_name_list (list): List of model names for our understanding
        saved_result_dir (str): Direcoty in which results are saved
    """

    if not os.path.exists(save_nn_dir):
        os.makedirs(save_nn_dir)
    for analyze_name, model_name in zip(analyze_name_list, model_name_list):
        saved_result_analysis_dir = os.path.join(saved_result_dir, analyze_name)
        for mode in ['people', 'scene']:
            if 'CAD' in analyze_name:
                mode = 'people'
            else:
                if not 'ours' in analyze_name:
                    mode = 'people'
                else:
                    mode = 'scene'

            nn_file_path = os.path.join(saved_result_analysis_dir, f'nn_video_{mode}.xlsx')
            nn_save_path = os.path.join(save_nn_dir, f'nn_video_{model_name}_{mode}.xlsx')
            shutil.copy(nn_file_path, nn_save_path)

def count_parameters(model):
    return sum(p.numel() for p in model.parameters() if p.requires_grad)

def get_activity_weight(query_stats_dir, use_query_type_list, dataset_name):
    activity_weight = {}
    for use_query_type in use_query_type_list:
        if dataset_name in ['vol', 'VOL']:
            query_stats_path = os.path.join(query_stats_dir, f'VOL_{use_query_type}_gt.json')
        elif dataset_name in ['cad', 'CAD']:
            query_stats_path = os.path.join(query_stats_dir, f'CAD_{use_query_type}_gt.json')
        elif dataset_name in ['bsk', 'BSK']:
            query_stats_path = os.path.join(query_stats_dir, f'BSK_{use_query_type}_gt.json')
        else:
            assert False, f'Unknown dataset name: {dataset_name}'

        with open(query_stats_path, 'r') as f:
            query_stats = json.load(f)
        query_stats_test = query_stats['test']
        query_stats_test_sum = query_stats_test['sum']
        activity_weight[use_query_type] = query_stats_test_sum
    
    # normalize the activity weight
    activity_weight_sum = sum(list(activity_weight.values()))
    for use_query_type in use_query_type_list:
        activity_weight[use_query_type] /= activity_weight_sum
    return activity_weight


def refine_config(cfg):
    if 'use_recon_loss' in dir(cfg):
        pass
    else:
        cfg.use_recon_loss = False

    if 'use_ind_feat' in dir(cfg):
        pass
    else:
        cfg.use_ind_feat = 'loc_and_app'

    if 'use_ind_feat_crop' in dir(cfg):
        pass
    else:
        cfg.use_ind_feat_crop = 'roi_multi'

    if 'person_size' in dir(cfg):
        pass
    else:
        cfg.person_size = 224, 224

    if 'trans_head_num' in dir(cfg):
        pass
    else:
        cfg.trans_head_num = 1

    if 'trans_layer_num' in dir(cfg):
        pass
    else:
        cfg.trans_layer_num = 1

    if 'final_head_mid_num' in dir(cfg):
        pass
    else:
        cfg.final_head_mid_num = 2

    return cfg